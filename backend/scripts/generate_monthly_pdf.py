import sys
import os
import json
import openpyxl
import subprocess

def main():
    if len(sys.argv) < 2:
        print("Error: JSON data path argument is required", file=sys.stderr)
        sys.exit(1)
        
    json_path = sys.argv[1]
    
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        period = data['period']  # e.g., '2026-06'
        shops = data['shops']
        
        # Get workspace root dynamically
        script_dir = os.path.dirname(os.path.abspath(__file__))
        workspace_root = os.path.abspath(os.path.join(script_dir, '../..'))
        
        # Load template
        template_path = os.path.join(workspace_root, 'meter-billing-converter/assets/meter_billing_template.xlsx')
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Parse year/month for output filename
        # e.g. '2026-06' -> '26年6月抄表'
        match = subprocess.run(['date'], capture_output=True) # unused, we can just parse period
        parts = period.split('-')
        year_short = parts[0][2:] if len(parts[0]) == 4 else parts[0]
        month = str(int(parts[1]))
        output_prefix = f"水电表{year_short}年{month}月抄表"
        
        # Let's map existing shops in the excel sheet by row index
        existing_shops = {}
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val:
                existing_shops[str(cell_val).strip()] = row_idx
                
        # Fill in the data
        for s in shops:
            shop_code = s['shop_code']
            shop_name = s['shop_name']
            labor_fee = s['labor_fee']
            rubbish_fee = s['rubbish_fee']
            
            # Find or append row
            row_idx = existing_shops.get(shop_code)
            if not row_idx:
                row_idx = ws.max_row + 1
                ws.cell(row=row_idx, column=1).value = shop_code
                existing_shops[shop_code] = row_idx
                
            ws.cell(row=row_idx, column=2).value = shop_name
            ws.cell(row=row_idx, column=13).value = labor_fee
            ws.cell(row=row_idx, column=14).value = rubbish_fee
            
            # Default prices
            water_price = 4.13
            elec_price = 1.03
            
            # Initialize meters values in row to None/0 first
            # Electic 1: Col 3, 4
            # Water: Col 5, 6
            # Electic 2: Col 7, 8
            # Electic 3: Col 9, 10
            ws.cell(row=row_idx, column=3).value = None
            ws.cell(row=row_idx, column=4).value = None
            ws.cell(row=row_idx, column=5).value = None
            ws.cell(row=row_idx, column=6).value = None
            ws.cell(row=row_idx, column=7).value = None
            ws.cell(row=row_idx, column=8).value = None
            ws.cell(row=row_idx, column=9).value = None
            ws.cell(row=row_idx, column=10).value = None
            
            for m in s['meters']:
                m_type = m['meter_type']
                m_name = m['meter_name']
                prev = m['previous_reading']
                curr = m['current_reading']
                price = m['unit_price']
                
                if m_type == 'water':
                    water_price = price
                    ws.cell(row=row_idx, column=5).value = prev
                    ws.cell(row=row_idx, column=6).value = curr
                elif m_type == 'electricity':
                    elec_price = price
                    if m_name == '电表1':
                        ws.cell(row=row_idx, column=3).value = prev
                        ws.cell(row=row_idx, column=4).value = curr
                    elif m_name == '电表2':
                        ws.cell(row=row_idx, column=7).value = prev
                        ws.cell(row=row_idx, column=8).value = curr
                    elif m_name == '电表3':
                        ws.cell(row=row_idx, column=9).value = prev
                        ws.cell(row=row_idx, column=10).value = curr
                        
            ws.cell(row=row_idx, column=11).value = water_price
            ws.cell(row=row_idx, column=12).value = elec_price
            
        # Save temp Excel file
        temp_xlsx = path = os.path.join(os.getcwd(), f"{output_prefix}.xlsx")
        wb.save(temp_xlsx)
        
        # Run generate_billing_pdf.py using same python env
        python_bin = sys.executable
        pdf_script = os.path.join(workspace_root, 'meter-billing-converter/scripts/generate_billing_pdf.py')
        output_pdf = os.path.join(os.getcwd(), f"{parts[0]}年{month}月抄表计费通知单.pdf")
        
        # Run subprocess
        result = subprocess.run(
            [python_bin, pdf_script, temp_xlsx, output_pdf],
            capture_output=True,
            text=True
        )
        
        # Clean up temp Excel
        try: os.remove(temp_xlsx)
        except: pass
        
        if result.returncode != 0:
            print(f"Error in generate_billing_pdf.py:\nStdout: {result.stdout}\nStderr: {result.stderr}", file=sys.stderr)
            sys.exit(1)
            
        # Return PDF path on stdout
        print(output_pdf)
        
    except Exception as e:
        print(f"Error: {str(e)}", file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    main()
