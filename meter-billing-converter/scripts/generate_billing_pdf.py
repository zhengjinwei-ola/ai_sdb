#!/usr/bin/env python3
import os
import re
import sys
import openpyxl

def to_chinese_uppercase(num):
    units = ['', '拾', '佰', '仟', '万', '拾', '佰', '仟', '亿']
    digits = ['零', '壹', '贰', '叁', '肆', '伍', '陆', '柒', '捌', '玖']
    num_str = f"{num:.2f}"
    integer_part, decimal_part = num_str.split('.')
    integer_val = int(integer_part)
    if integer_val == 0:
        out_int = '零'
    else:
        out_int = ''
        n_str = str(integer_val)
        length = len(n_str)
        for i, d in enumerate(n_str):
            digit = int(d)
            pos = length - 1 - i
            unit = units[pos]
            if digit != 0:
                out_int += digits[digit] + unit
            else:
                if out_int and not out_int.endswith('零') and pos % 4 != 0:
                    out_int += '零'
                if pos == 4 and not out_int.endswith('万'):
                    if out_int.endswith('零'):
                        out_int = out_int[:-1] + '万'
                    else:
                        out_int += '万'
        if out_int.endswith('零'):
            out_int = out_int[:-1]
    dec_val = int(decimal_part)
    jiao = dec_val // 10
    fen = dec_val % 10
    out_dec = ''
    if jiao == 0 and fen == 0:
        out_dec = '元整'
    else:
        out_dec = '元'
        if jiao != 0:
            out_dec += digits[jiao] + '角'
        else:
            out_dec += '零'
        if fen != 0:
            out_dec += digits[fen] + '分'
    result = out_int + out_dec
    if result.startswith('元'):
        result = result[1:]
    return result

def round_half_up(val):
    return int(val + 0.5)

def parse_float(val):
    if val is None or str(val).strip() == '':
        return 0.0
    try:
        return float(val)
    except ValueError:
        return 0.0

def parse_int(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        return int(float(val))
    except ValueError:
        return None

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_billing_pdf.py <excel_file> [output_pdf]")
        sys.exit(1)
        
    excel_path = sys.argv[1]
    filename = os.path.basename(excel_path)
    
    # Extract year and month from filename
    match = re.search(r'水电表(?:20)?(\d+)年(\d+)月抄表', filename)
    if match:
        year = int(match.group(1))
        if year < 100:
            year += 2000
        month = int(match.group(2))
    else:
        import datetime
        now = datetime.datetime.now()
        year = now.year
        month = now.month
        
    # Output path
    if len(sys.argv) >= 3:
        pdf_path = sys.argv[2]
    else:
        pdf_path = f"{year}年{month}月抄表计费通知单.pdf"
        
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = wb.active
    
    # Process rows
    shops = []
    for r_idx in range(2, sheet.max_row + 1):
        row = [sheet.cell(r_idx, col_idx).value for col_idx in range(1, 15)]
        shop_id = row[0]
        shop_name = row[1]
        
        # Skip empty rows
        if shop_id is None or str(shop_id).strip() == '' or shop_name is None or str(shop_name).strip() == '':
            continue
            
        shop_id = str(shop_id).strip()
        shop_name = str(shop_name).strip()
        
        # Parse electric meters
        meters = []
        m1_prev = parse_int(row[2])
        m1_curr = parse_int(row[3])
        
        m2_prev = parse_int(row[6])
        m2_curr = parse_int(row[7])
        
        m3_prev = parse_int(row[8])
        m3_curr = parse_int(row[9])
        
        has_multi = (m2_prev is not None or m2_curr is not None or m3_prev is not None or m3_curr is not None)
        
        if m1_prev is not None or m1_curr is not None:
            meters.append({
                'name': '电表 1' if has_multi else '电表',
                'prev': m1_prev or 0,
                'curr': m1_curr or 0,
            })
            
        if m2_prev is not None or m2_curr is not None:
            meters.append({
                'name': '电表 2',
                'prev': m2_prev or 0,
                'curr': m2_curr or 0,
            })
            
        if m3_prev is not None or m3_curr is not None:
            meters.append({
                'name': '电表 3',
                'prev': m3_prev or 0,
                'curr': m3_curr or 0,
            })
            
        # Water meter
        water_prev = parse_int(row[4]) or 0
        water_curr = parse_int(row[5]) or 0
        water_diff = max(0, water_curr - water_prev)
        
        water_price = parse_float(row[10])
        elec_price = parse_float(row[11])
        labor_fee = parse_float(row[12])
        garbage_fee = parse_float(row[13])
        
        # Calculate fees
        total_elec_diff = sum(max(0, m['curr'] - m['prev']) for m in meters)
        elec_amount = round_half_up(total_elec_diff * elec_price)
        water_amount = round_half_up(water_diff * water_price)
        
        subtotal = elec_amount + water_amount
        total = subtotal + labor_fee + garbage_fee
        
        uppercase_total = to_chinese_uppercase(total)
        
        shops.append({
            'id': shop_id,
            'name': shop_name,
            'meters': meters,
            'water': {
                'prev': water_prev,
                'curr': water_curr,
                'diff': water_diff,
                'price': water_price,
                'amount': water_amount
            },
            'elec_price': elec_price,
            'elec_amount': elec_amount,
            'labor_fee': labor_fee,
            'garbage_fee': garbage_fee,
            'subtotal': subtotal,
            'total': total,
            'uppercase_total': uppercase_total
        })

    # Format Date
    # Match the double space for single-digit month
    date_month_str = f"  {month}" if month < 10 else f" {month}"
    date_str = f"{year} 年{date_month_str} 月 15 日"

    # Style definitions
    font_family = "font-family: SimSun, 'PingFang SC', 'Microsoft YaHei', sans-serif;"
    
    # Generate HTML content
    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>{year}年{month}月抄表计费通知单</title>
<style>
@page {{
    size: A4;
    margin: 8mm 15mm 8mm 15mm;
}}
body {{
    margin: 0;
    padding: 0;
    background-color: #fff;
    -webkit-print-color-adjust: exact;
}}
</style>
</head>
<body>
"""

    # Add notices
    for i, shop in enumerate(shops):
        # Apply page break before every 3rd notice (index 3, 6, 9...)
        if i > 0 and i % 3 == 0:
            html += f'  <p style="page-break-before: always; margin: 0; padding: 0; line-height: 0;">&nbsp;</p>\n'
        
        html += f'  <div class="notice" style="box-sizing: border-box; padding: 1px 0; margin-bottom: 1.5mm;">\n'
        html += f'    <div class="title" style="font-size: 12.5pt; font-weight: bold; text-align: center; margin-bottom: 1mm; letter-spacing: 1px; {font_family}">{year} 年 {month} 月 抄表计费通知单</div>\n'
        
        # Meta-table with explicit inline styles
        html += f'    <table style="width: 100%; border-collapse: collapse; border: none; margin-bottom: 1mm; font-size: 8.5pt; {font_family}">\n'
        html += f'      <tr>\n'
        html += f'        <td style="width: 25%; border: none; padding: 1px 0;">编号： {shop["id"]}</td>\n'
        html += f'        <td style="width: 25%; border: none; padding: 1px 0;">姓名 {shop["name"]}</td>\n'
        html += f'        <td style="width: 20%; border: none; padding: 1px 0;">抄表人： 李</td>\n'
        html += f'        <td style="width: 30%; border: none; padding: 1px 0;">抄表日期：{date_str}</td>\n'
        html += f'      </tr>\n'
        html += f'    </table>\n'
        
        # Data-table with explicit inline styles for ALL borders and paddings
        html += f'    <table style="width: 100%; border-collapse: collapse; border: 1px solid #000000; font-size: 8.5pt; text-align: center; margin-bottom: 1mm; {font_family}">\n'
        html += f'      <thead>\n'
        html += f'        <tr>\n'
        html += f'          <th style="width: 18%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">项目</th>\n'
        html += f'          <th style="width: 14%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">上月表底</th>\n'
        html += f'          <th style="width: 14%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">本月抄表数</th>\n'
        html += f'          <th style="width: 13%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">实用度数</th>\n'
        html += f'          <th style="width: 13%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">公共分摊</th>\n'
        html += f'          <th style="width: 14%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">单价（元）</th>\n'
        html += f'          <th style="width: 14%; border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2;">金额</th>\n'
        html += f'        </tr>\n'
        html += f'      </thead>\n'
        html += f'      <tbody>\n'
        
        # Meter rows
        for idx, m in enumerate(shop['meters']):
            diff = max(0, m['curr'] - m['prev'])
            if idx == 0:
                html += f'        <tr>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["name"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["prev"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["curr"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{diff}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["elec_price"]:.2f}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["elec_amount"]}</td>\n'
                html += f'        </tr>\n'
            else:
                html += f'        <tr>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["name"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["prev"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{m["curr"]}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{diff}</td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
                html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
                html += f'        </tr>\n'
                
        # Water row
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">水费</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["water"]["prev"]}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["water"]["curr"]}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["water"]["diff"]}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["water"]["price"]:.3f}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["water"]["amount"]}</td>\n'
        html += f'        </tr>\n'
        
        # Other rows
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">水电人工<br>费</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["labor_fee"]:.2f}</td>\n'
        html += f'        </tr>\n'
        
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">垃圾处理<br>费</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">{shop["garbage_fee"]:.2f}</td>\n'
        html += f'        </tr>\n'
        
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">滞纳金</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">0.00</td>\n'
        html += f'        </tr>\n'
        
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">广告费</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;"></td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 2.2px 2px; height: 13px; vertical-align: middle;">0.00</td>\n'
        html += f'        </tr>\n'
        
        # Total row with inline style
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 3px 2px; font-weight: bold; background-color: #f2f2f2; height: 14px; vertical-align: middle;">合计</td>\n'
        html += f'          <td colspan="6" style="border: 1px solid #000000; padding: 3px 2px; text-align: left; padding-left: 8px; font-weight: bold; font-size: 8.5pt; height: 14px; vertical-align: middle;">大写：{shop["uppercase_total"]} &nbsp;&nbsp;&nbsp;&nbsp; 小写： {shop["total"]:.2f}</td>\n'
        html += f'        </tr>\n'
        
        html += f'      </tbody>\n'
        html += f'    </table>\n'
        
        html += f'    <div class="footer-notes" style="font-size: 7pt; line-height: 1.2; margin-top: 0.3mm; {font_family}">\n'
        html += f'      1、此单可对账不做凭证；2、每月5日前为收费时间，超期按5%收滞纳金 or 停电；3、以上费用如有不明或差请到管理处核对。\n'
        html += f'    </div>\n'
        html += f'  </div>\n'
        
        # Add divider line only between notices on the SAME page
        pos_in_page = i % 3
        if pos_in_page < 2 and i < len(shops) - 1:
            html += f'  <div class="divider-line" style="text-align: center; font-size: 9pt; margin: 1.5mm 0; border-top: 1px dashed #000; height: 1px;">========================================</div>\n'

    # Summary Page (always breaks page dynamically via CSS page-break-before)
    html += f'  <p style="page-break-before: always; margin: 0; padding: 0; line-height: 0;">&nbsp;</p>\n'
    html += f'  <div class="summary-page" style="page-break-inside: avoid; box-sizing: border-box; padding-top: 5mm;">\n'
    html += f'    <div class="summary-title" style="font-size: 18pt; font-weight: bold; text-align: center; margin-bottom: 10mm; {font_family}">费用汇总表</div>\n'
    
    # Summary Table with explicit inline styles
    html += f'    <table class="summary-table" style="width: 100%; border-collapse: collapse; border: 1px solid #000000; font-size: 10.5pt; text-align: center; {font_family}">\n'
    html += f'      <thead>\n'
    html += f'        <tr>\n'
    html += f'          <th style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">店铺名称</th>\n'
    html += f'          <th style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">水电费合计<br>（元）</th>\n'
    html += f'          <th style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">水电人工费</th>\n'
    html += f'          <th style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">垃圾处理费</th>\n'
    html += f'          <th style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">总价</th>\n'
    html += f'        </tr>\n'
    html += f'      </thead>\n'
    html += f'      <tbody>\n'
    
    total_subtotal = 0.0
    total_labor = 0.0
    total_garbage = 0.0
    total_total = 0.0
    
    for shop in shops:
        html += f'        <tr>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 6px 4px;">{shop["name"]}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 6px 4px;">{shop["subtotal"]:.2f}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 6px 4px;">{shop["labor_fee"]:.2f}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 6px 4px;">{shop["garbage_fee"]:.2f}</td>\n'
        html += f'          <td style="border: 1px solid #000000; padding: 6px 4px;">{shop["total"]:.2f}</td>\n'
        html += f'        </tr>\n'
        
        total_subtotal += shop['subtotal']
        total_labor += shop['labor_fee']
        total_garbage += shop['garbage_fee']
        total_total += shop['total']
        
    html += f'        <tr style="font-weight: bold;">\n'
    html += f'          <td style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">合计</td>\n'
    html += f'          <td style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">{total_subtotal:.2f}</td>\n'
    html += f'          <td style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">{total_labor:.2f}</td>\n'
    html += f'          <td style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">{total_garbage:.2f}</td>\n'
    html += f'          <td style="border: 1px solid #000000; padding: 6px 4px; font-weight: bold; background-color: #f2f2f2;">{total_total:.2f}</td>\n'
    html += f'        </tr>\n'
    
    html += f'      </tbody>\n'
    html += f'    </table>\n'
    html += f'  </div>\n'

    html += """</body>
</html>
"""

    # Write HTML file
    temp_html_path = "temp_billing.html"
    with open(temp_html_path, 'w', encoding='utf-8') as f:
        f.write(html)
        
    # Convert to PDF using soffice (force standard Writer mode for proper paginated A4 layout)
    soffice_path = "/opt/homebrew/bin/soffice"
    if not os.path.exists(soffice_path):
        print(f"Error: LibreOffice '{soffice_path}' not found.")
        sys.exit(1)
        
    # Standard HTML (StarWriter) filter enforces page-break properties correctly into Writer A4
    cmd = f"{soffice_path} --headless --infilter='HTML (StarWriter)' --convert-to pdf {temp_html_path}"
    ret = os.system(cmd)
    if ret != 0:
        # Fallback to standard conversion
        cmd = f"{soffice_path} --headless --convert-to pdf {temp_html_path}"
        ret = os.system(cmd)
        if ret != 0:
            print("Error converting HTML to PDF using LibreOffice.")
            sys.exit(1)
        
    # Rename output PDF
    if os.path.exists("temp_billing.pdf"):
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        os.rename("temp_billing.pdf", pdf_path)
        print(f"Success! Generated '{pdf_path}' from '{excel_path}'")
    else:
        print("Error: temp_billing.pdf was not generated.")
        sys.exit(1)
        
    # Clean up temp HTML
    if os.path.exists(temp_html_path):
        os.remove(temp_html_path)

if __name__ == '__main__':
    main()
